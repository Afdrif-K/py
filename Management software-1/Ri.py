#encoding=utf-8
from openpyxl import load_workbook
import pandas as pd
from nameout import out

R=input("输入年份:")
date=pd.date_range(R+'/01/01',R+'/12/31', freq='D')
print(date)
week=[int(i.strftime("%w")) for i in date]  # 0表示星期日
dataframe = pd.DataFrame({'date':date,'week':week})
dataframe.to_excel('dates.xlsx',index=False)
#------------------------------------------------------execl
wb=load_workbook('./dates.xlsx')
# print(wb.sheetnames)
sheet=wb["Sheet1"]
sheet.title='date'
# 对行进行遍历,输出A1,B1,C1
# for row in sheet.rows:
#     for cell in row:
#         print(cell.value)
# 对列进行遍历,输出A1,A2,A3
# for column in sheet.columns:
#     for cell in column:
#         print(cell.value)
for column in sheet['A']:
	column.value=str(column.value).replace(' 00:00:00','')
	# print(column.value)

for column in sheet['B']:
	# print(column.value)
	if column.value ==  1:
		column.value='星期一'
		# print(column.value)
	elif column.value ==2:
		column.value = '星期二'
	elif column.value ==3:
		column.value = '星期三'
	elif column.value ==4:
		column.value = '星期四'
	elif column.value ==5:
		column.value = '星期五'
	elif column.value ==6:
		column.value = '星期六'
	elif column.value ==0:
		column.value = '星期日'
wb.save('./dates.xls')
out()
