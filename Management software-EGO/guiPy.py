#encoding=utf-8
import tkinter as tk
from tkinter import ttk
from dataBase import *
import tkinter.messagebox
from tkinter.ttk import *
import xlwt
import xlrd
from time import sleep
import pymysql
import win32com.client as win32
import os
from openpyxl import load_workbook


class MainWindow(tk.Tk):
    def __init__(self):
        super().__init__()
        # 变量定义
        self.opr = mysql()
        self.list = self.init_data()
        # print(self.list)
        self.item_selection = ''
        self.item_selections = ''
        self.data = []
        self.datas = []

        # 定义区域，把全局分为上中下三部分
        self.frame_top = tk.Frame(width=1000, height=90)
        self.frame_center = tk.Frame(width=1000, height=330)
        self.frame_bottom = tk.Frame(width=1000, height=250)

        # 定义上部分区域
        self.lb_tips = tk.Label(self.frame_top, text="                                                          ")
        self.lb_tip = tk.Label(self.frame_top, text="Class名称:")
        self.string = tk.StringVar()
        self.string.set('')
        self.ent_find_name = tk.Entry(self.frame_top, textvariable=self.string)
        self.btn_query = tk.Button(self.frame_top, text="查询", command=self.query)
        self.btn_querys = tk.Button(self.frame_top, text="刷新", command=self.shu)
        self.lb_tips.grid(row=0,column=3)
        self.lb_tip.grid(row=0, column=10,padx=15,pady=35)
        self.ent_find_name.grid(row=0, column=11, padx=10, pady=35)
        self.btn_query.grid(row=0, column=12, padx=20, pady=35)
        self.btn_querys.grid(row=0, column=13, padx=15, pady=35)

        # 定义下部分区域
        self.btn_delete = tk.Button(self.frame_bottom, text="删除", command=self.delete)
        self.btn_update = tk.Button(self.frame_bottom, text="修改", command=self.update)
        self.btn_add = tk.Button(self.frame_bottom, text="添加", command=self.add)
        self.btn_get = tk.Button(self.frame_bottom, text="考勤管理", command=self.gets)
        self.btn_delete.grid(row=0, column=0, padx=80, pady=45)
        self.btn_update.grid(row=0, column=1, padx=80, pady=45)
        self.btn_add.grid(row=0, column=2, padx=80, pady=45)
        self.btn_get.grid(row=0, column=3, padx=80, pady=45)


        # 定义中心列表区域
        self.tree = ttk.Treeview(self.frame_center, show="headings", height=15, columns=("a", "b", "c", "d","e","f","g","h"))
        self.vbar = ttk.Scrollbar(self.frame_center, orient=tk.VERTICAL, command=self.tree.yview)
        self.vbars = ttk.Scrollbar(self.frame_center, orient=tk.HORIZONTAL, command=self.tree.xview)
        # 定义树形结构与滚动条
        self.tree.configure(xscrollcommand=self.vbars.set)
        self.tree.configure(yscrollcommand=self.vbar.set)
        # 表格的标题
        self.tree.column("a", width=30, anchor="center")
        self.tree.column("b", width=85, anchor="center")
        self.tree.column("c", width=85, anchor="center")
        self.tree.column("d", width=80, anchor="center")
        self.tree.column("e", width=220, anchor="center")
        self.tree.column("f", width=90, anchor="center")
        self.tree.column("g", width=60, anchor="center")
        self.tree.column("h", width=220, anchor="center")
        self.tree.heading("a", text="编号")
        self.tree.heading("b", text="学生名称")
        self.tree.heading("c", text="负责老师")
        self.tree.heading("d", text="班级")
        self.tree.heading("e", text="上课时间")
        self.tree.heading("f", text="加入时间")
        self.tree.heading("g", text="赠送课时")
        self.tree.heading("h", text="赠送课时信息记录")
        # 调用方法获取表格内容插入及树基本属性设置
        self.tree["selectmode"] = "browse"
        self.get_tree()
        self.tree.grid(row=0, column=0, sticky=tk.NSEW, ipadx=10)
        self.vbar.grid(row=0, column=1, sticky=tk.NS)
        self.vbars.grid(row=1, column=0, sticky=tk.EW)


        # 定义整体区域
        self.frame_top.grid(row=0, column=0, padx=60)
        self.frame_center.grid(row=2, column=0, padx=60, ipady=1)
        self.frame_bottom.grid(row=4, column=0, padx=60)
        self.frame_top.grid_propagate(0)
        self.frame_center.grid_propagate(0)
        self.frame_bottom.grid_propagate(0)

        # 窗体设置
        self.center_window(1000 ,550)
        self.title('人员管理')
        self.resizable(False, False)
        self.mainloop()

    # 窗体居中
    def center_window(self, width, height):
        screenwidth = self.winfo_screenwidth()
        screenheight = self.winfo_screenheight()
        # 宽高及宽高的初始点坐标
        size = '%dx%d+%d+%d' % (width, height, (screenwidth - width) / 2, (screenheight - height) / 2)
        self.geometry(size)

    # 数据初始化获取
    def init_data(self):
        result, _ = self.opr.query()
        # print(result)
        if self.opr.queryStatus:
            return 0
        else:
            return result

    # 表格内容插入
    def get_tree(self):
        if self.list == 0:
            tkinter.messagebox.showinfo("错误提示", "数据获取失败")
        else:
            # 删除原节点
            for _ in map(self.tree.delete, self.tree.get_children("")):
                pass
            # 更新插入新节点
            for i in range(len(self.list)):
                group = self.list[i]
                self.tree.insert("", "end", values=(group[0],
                                                    group[1],
                                                    group[2],
                                                    group[3],
                                                    group[4],
                                                    group[5],
                                                    group[6],
                                                    group[7]), text=group[0])
            # TODO 此处需解决因主程序自动刷新引起的列表项选中后重置的情况，我采用的折中方法是：把选中时的数据保存下来，作为记录
            # 绑定列表项单击事件
            self.tree.bind("<ButtonRelease-1>", self.tree_item_click)
            self.tree.after(200000, self.get_tree)

    # 单击查询按钮触发的事件方法
    def query(self):
        query_info = self.ent_find_name.get()
        # print(query_info)
        self.string.set('')
        # print(query_info)
        if query_info is None or query_info == '':
            tkinter.messagebox.showinfo("警告", "查询条件不能为空！")
            self.get_tree()
        else:
            result, _ = self.opr.query(queryby="where Class like '" + query_info + "';")
            if self.opr.queryStatus:
                tkinter.messagebox.showinfo("警告", "查询出错，请检查数据库服务是否正常")
            elif not result:
                tkinter.messagebox.showinfo("查询结果", "该查询条件没有匹配项！")
            else:
                self.list = result
                print(self.list)
                self.get_tree()

    def shu(self):
        self.list = self.init_data()
        self.get_tree()

    # 为解决窗体自动刷新的问题，记录下单击项的内容
    def tree_item_click(self, event):
        try:
            selection = self.tree.selection()[0]
            self.data = self.tree.item(selection, "values")
            self.datas = self.tree.item(selection, "values")
            # print(self.datas)
            self.item_selection = self.data[0]
            self.item_selections = self.data[1]
            # print(self.item_selections)
        except IndexError:
            tkinter.messagebox.showinfo("单击警告", "单击结果范围异常，请重新选择！")

    # 单击删除按钮触发的事件方法
    def delete(self):
        if self.item_selection is None or self.item_selection == '':
            tkinter.messagebox.showinfo("删除警告", "未选中待删除值")
        else:
            # TODO： 删除提示
            a=tkinter.messagebox.askokcancel('删除警告', '要执行此操作吗')
            if a==True:
                self.opr.delete(deleteby="id ="+self.item_selection,name=self.item_selections)
            # deleteby = "id = " + self.item_selection
            # print(deleteby)
            # print(self.opr.deleteStatus)
            if self.opr.deleteStatus:
                tkinter.messagebox.showinfo("删除警告", "删除异常，可能是数据库服务意外关闭了。。。")
            else:
                self.list = self.init_data()
                self.get_tree()

    # 单击更新按钮触发的事件方法
    def update(self):
        if self.item_selection is None or self.item_selection == '':
            tkinter.messagebox.showinfo("更新警告", "未选中待更新项")
        else:
            data = [self.item_selection]
            self.data = self.set_info(2)
            # print(self.data)
            if self.data is None or not self.data:
                return
            # 更改参数
            data = data + self.data
            # print(data)
            self.opr.update(updatelist=data)
            # print(self.opr.updateStatus)
            if self.opr.updateStatus:
                tkinter.messagebox.showinfo("更新信息警告", "数据异常库连接异常，可能是服务关闭啦~")
            # 更新界面，刷新数据
            self.list = self.init_data()
            self.get_tree()

    # 单击新增按钮触发的事件方法
    def add(self):
        # 接收弹窗的数据
        self.data = self.set_info(1)
        # print(self.data)
        if self.data is None or not self.data:
            return
        # 更改参数
        self.opr.insert(insertlist=self.data)
        # print(self.opr.insertStatus)
        if self.opr.insertStatus:
            tkinter.messagebox.showinfo("新增信息警告", "数据异常库连接异常，可能是服务关闭啦~")
        # 更新界面，刷新数据
        self.list = self.init_data()
        self.get_tree()

    def gets(self):
        if self.item_selections is None or self.item_selections == '':
            tkinter.messagebox.showinfo("警告", "未选中待考勤项")
        else:
            datas = [self.item_selections]
            self.datas = self.set_infos(3)
            if self.datas is None or not self.datas:
                return

    def set_infos(self, dia_type):
        """
        :param dia_type:表示打开的是新增窗口还是更新窗口，新增则参数为1，其余参数为更新
        :return: 返回用户填写的数据内容，出现异常则为None
        """
        dialogs = gett(datas=self.datas, dia_types=dia_type)
        # self.withdraw()
        self.wait_window(dialogs)  # 这一句很重要！！！
        return dialogs

    # 此方法调用弹窗传递参数，并返回弹窗的结果
    def set_info(self, dia_type):
        """
        :param dia_type:表示打开的是新增窗口还是更新窗口，新增则参数为1，其余参数为更新
        :return: 返回用户填写的数据内容，出现异常则为None
        """
        dialog = MyDialog(data=self.data, dia_type=dia_type)
        # self.withdraw()
        self.wait_window(dialog)  # 这一句很重要！！！
        return dialog.group_info


# 新增窗口或者更新窗口
class MyDialog(tk.Toplevel):
    def __init__(self, data, dia_type):
        # print(data)
        super().__init__()
        global a
        a=dia_type
        # 窗口初始化设置，设置大小，置顶等
        self.center_window(700, 560)
        self.wm_attributes("-topmost", 1)
        self.resizable(False, False)
        self.protocol("WM_DELETE_WINDOW", self.donothing)   # 此语句用于捕获关闭窗口事件，用一个空方法禁止其窗口关闭。

        # 根据参数类别进行初始化
        if dia_type == 1:
            self.title('新增信息')
        else:
            self.title('更新信息')

        # 数据变量定义
        self.id = tk.StringVar()
        self.name = tk.StringVar()
        self.Tname = tk.StringVar()
        self.Class = tk.StringVar()
        self.time = tk.StringVar()
        self.times = tk.StringVar()
        self.ke = tk.StringVar()
        self.kes = tk.StringVar()
        if not data or dia_type == 1:
            self.id.set('')
            self.name.set('')
            self.Tname.set('')
            self.Class.set('')
            self.time.set('')
            self.times.set('')
            self.ke.set('')
            self.kes.set('')

        else:
            self.id.set(data[0])
            self.name.set(data[1])
            self.Tname.set(data[2])
            self.Class.set(data[3])
            self.time.set(data[4])
            self.times.set(data[5])
            self.ke.set(data[6])
            self.kes.set(data[7])

        # 错误提示定义
        self.text_error_id = tk.StringVar()
        self.text_error_name = tk.StringVar()
        self.text_error_Tname = tk.StringVar()
        self.text_error_Class = tk.StringVar()
        self.text_error_time = tk.StringVar()
        self.text_error_times = tk.StringVar()
        self.text_error_ke = tk.StringVar()
        self.text_error_kes = tk.StringVar()
        self.error_null = '该项内容不能为空!'
        self.error_exsit = '该编号已存在!'
        self.error_isdigit = '名字不能含有数字'

        self.group_info = []
        # 弹窗界面布局
        self.setup_ui()

    # 窗体布局设置
    def setup_ui(self):
        # 第一行（两列）
        row1 = tk.Frame(self)
        row1.grid(row=0, column=0, padx=160, pady=20)
        tk.Label(row1, text='编号：', width=8).pack(side=tk.LEFT)
        tk.Entry(row1, textvariable=self.id, width=20).pack(side=tk.LEFT)
        tk.Label(row1, textvariable=self.text_error_id, width=20, fg='red').pack(side=tk.LEFT)
        # 第二行
        row2 = tk.Frame(self)
        row2.grid(row=1, column=0, padx=160, pady=20)
        tk.Label(row2, text='学生名称：', width=8).pack(side=tk.LEFT)
        tk.Entry(row2, textvariable=self.name, width=20).pack(side=tk.LEFT)
        tk.Label(row2, textvariable=self.text_error_name, width=20, fg='red').pack(side=tk.LEFT)
        # 第三行
        row3 = tk.Frame(self)
        row3.grid(row=2, column=0, padx=160, pady=20)
        tk.Label(row3, text='责任老师：', width=10).pack(side=tk.LEFT)
        tk.Entry(row3, textvariable=self.Tname, width=18).pack(side=tk.LEFT)
        tk.Label(row3, textvariable=self.text_error_Tname, width=20, fg='red').pack(side=tk.LEFT)
        # 第四行
        row4 = tk.Frame(self)
        row4.grid(row=3, column=0, padx=160, pady=20)
        tk.Label(row4, text='班级：', width=8).pack(side=tk.LEFT)
        tk.Entry(row4, textvariable=self.Class, width=20).pack(side=tk.LEFT)
        tk.Label(row4, textvariable=self.text_error_Class, width=20, fg='red').pack(side=tk.LEFT)
        #
        row5 = tk.Frame(self)
        row5.grid(row=4, column=0, padx=160, pady=20)
        tk.Label(row5, text='上课时间：', width=8).pack(side=tk.LEFT)
        tk.Entry(row5, textvariable=self.time, width=20).pack(side=tk.LEFT)
        tk.Label(row5, textvariable=self.text_error_time, width=20, fg='red').pack(side=tk.LEFT)
        #
        row6 = tk.Frame(self)
        row6.grid(row=5, column=0, padx=160, pady=20)
        tk.Label(row6, text='加入时间：', width=8).pack(side=tk.LEFT)
        tk.Entry(row6, textvariable=self.times, width=20).pack(side=tk.LEFT)
        tk.Label(row6, textvariable=self.text_error_times, width=20, fg='red').pack(side=tk.LEFT)
        # 第7行
        row7 = tk.Frame(self)
        row7.grid(row=6, column=0, padx=160, pady=20)
        tk.Label(row7, text='赠送课时：', width=8).pack(side=tk.LEFT)
        tk.Entry(row7, textvariable=self.ke, width=20).pack(side=tk.LEFT)
        tk.Label(row7, textvariable=self.text_error_ke, width=20, fg='red').pack(side=tk.LEFT)
        #
        row8 = tk.Frame(self)
        row8.grid(row=7, column=0, padx=160, pady=20)
        tk.Label(row8, text='赠送记录：', width=8).pack(side=tk.LEFT)
        tk.Entry(row8, textvariable=self.kes, width=20).pack(side=tk.LEFT)
        tk.Label(row8, textvariable=self.text_error_kes, width=20, fg='red').pack(side=tk.LEFT)
        #
        row9 = tk.Frame(self)
        row9.grid(row=8, column=0, padx=160, pady=5)
        tk.Button(row9, text="取消", command=self.cancel).grid(row=0, column=0, padx=60)
        tk.Button(row9, text="确定", command=self.ok).grid(row=0, column=1, padx=60)

    def center_window(self, width, height):
        screenwidth = self.winfo_screenwidth()
        screenheight = self.winfo_screenheight()
        size = '%dx%d+%d+%d' % (width, height, (screenwidth - width) / 2, (screenheight - height) / 2)
        self.geometry(size)

    # 点击确认按钮绑定事件方法
    def ok(self):
        self.group_info = [self.id.get(), self.name.get(), self.Tname.get(), self.Class.get(),self.time.get(),self.times.get(),self.ke.get(),self.kes.get()]  # 设置数据
        if self.check_info() == 1:  # 进行数据校验，失败则不关闭窗口
            return
        self.destroy()  # 销毁窗口

    # 点击取消按钮绑定事件方法
    def cancel(self):
        self.group_info = None  # 空！
        self.destroy()

    # 数据校验和用户友好性提示，校验失败返回1，成功返回0
    def check_info(self):
        is_null = 0
        str_tmp = self.group_info
        print(str_tmp[1].isdigit())
        if str_tmp[0] == '':
            self.text_error_id.set(self.error_null)
            is_null = 1
        if str_tmp[1] == '':
            self.text_error_name.set(self.error_null)
            is_null = 1
        if str_tmp[2] == '':
            self.text_error_Tname.set(self.error_null)
            is_null = 1
        if str_tmp[3] == '':
            self.text_error_Class.set(self.error_null)
            is_null = 1
        if str_tmp[4] == '':
            self.text_error_time.set(self.error_null)
            is_null = 1
        if str_tmp[1].isdigit() == True:
            self.text_error_name.set(self.error_isdigit)
            is_null = 1

        if is_null == 1:
            return 1
        res, _ = mysql().query(queryby="where id = '"+str_tmp[0]+"';")
        # print(res)
        if res:
            if a == 1:
                self.text_error_id.set(self.error_exsit)
                return 1
        return 0

    # 空函数
    def donothing(self):
        pass


from tkinter import *
class gett(tk.Tk):
    def __init__(self, datas, dia_types):
        super().__init__()
        # 窗口初始化设置，设置大小，置顶等
        self.center_windows(1000, 700)
        # self.wm_attributes("-topmost", 1)
        self.resizable(False, False)
        # self.protocol("WM_DELETE_WINDOW", self.donothings)   # 此语句用于捕获关闭窗口事件，用一个空方法禁止其窗口关闭。
        self.opr = mysql()
        self.list = self.init_data(datas)
        self.setup_uis(datas)
        self.s=datas

        # 窗体布局设置
    def setup_uis(self,datas):
        # 第一行（两列）
        self.row1 = tk.Frame(self)
        self.row1.grid(row=0, column=0, padx=10, pady=40)
        radVars = IntVar()
        self.f=tk.Button(self.row1,text="1",command=self.gosf)
        self.f.grid(row=0, column=0, padx=20)
        self.e=tk.Button(self.row1,text="2",command=self.gose)
        self.e.grid(row=1, column=0, padx=20)
        self.Q = tk.Button(self.row1, text="Q1",command=self.gosq)
        self.Q.grid(row=0, column=1, padx=20)
        self.Q = tk.Button(self.row1, text="Q2", command=self.gosqq)
        self.Q.grid(row=1, column=1, padx=20)
        self.S = tk.Button(self.row1, text="S1",command=self.goss)
        self.S.grid(row=0, column=2, padx=20)
        self.S = tk.Button(self.row1, text="S2", command=self.gosss)
        self.S.grid(row=1, column=2, padx=20)
        self.E = tk.Button(self.row1, text="空", command=self.gos0)
        self.E.grid(row=2, column=0, padx=20)
        Label(self.row1, text="年份:").grid(row=0, column=3, padx=60)
        Label(self.row1, text="月份:").grid(row=0, column=4, padx=60)
        self.C1=Combobox(self.row1,values=[x for x in range(2001, 2200)])
        self.C1.grid(row=1, column=3,padx=60)
        self.C2=Combobox(self.row1,values=[y for y in range(1, 13)])
        self.C2.grid(row=1, column=4, padx=60)
        self.B1=Button(self.row1, text="显示", command=self.go)
        self.B1.grid(row=1, column=5, padx=20)
        self.btn_queryss = tk.Button(self.row1, text="刷新", command=self.shu)
        self.btn_queryss.grid(row=1, column=6, padx=20)

        self.row2 = tk.Frame(self)
        self.row2.grid(row=1, column=0, padx=85, pady=15)
        # 定义中心列表区域
        self.tree = ttk.Treeview(self.row2, show="headings", height=20, columns=("a","b","c"))
        self.vbar = ttk.Scrollbar(self.row2, orient=tk.VERTICAL, command=self.tree.yview)
        self.vbars = ttk.Scrollbar(self.row2, orient=tk.HORIZONTAL, command=self.tree.xview)
        # 定义树形结构与滚动条
        self.tree.configure(xscrollcommand=self.vbars.set)
        self.tree.configure(yscrollcommand=self.vbar.set)
        # 表格的标题
        self.tree.column("a", width=200, anchor="center")
        self.tree.column("b", width=300, anchor="center")
        self.tree.column("c", width=300, anchor="center")
        self.tree.heading("a", text="date")
        self.tree.heading("b", text="week")
        self.tree.heading("c", text=datas[1])
        # 调用方法获取表格内容插入及树基本属性设置
        self.tree["selectmode"] = "browse"
        self.get_tree()
        self.tree.grid(row=0, column=0, sticky=tk.NSEW, ipadx=10)
        self.vbar.grid(row=0, column=1, sticky=tk.NS)
        self.vbars.grid(row=1, column=0, sticky=tk.EW)

        self.row3 = tk.Frame(self)
        self.row3.grid(row=2, column=0, padx=10, pady=1)
        self.btn_dao = tk.Button(self.row3, text="导出整体考勤数据表格", command=self.dao)
        self.btn_dao.grid(row=0, column=0, padx=10)


    def get_tree(self):
        if self.list == 0:
            tkinter.messagebox.showinfo("错误提示", "数据获取失败")
        else:
            for _ in map(self.tree.delete, self.tree.get_children("")):
                pass
            for i in range(len(self.list)):
                group = self.list[i]
                self.tree.insert("", "end", values=(group[0],
                                                    group[1],
                                                    group[2]), text=group[0])
            self.tree.bind("<ButtonRelease-1>", self.tree_item_click)
            self.tree.after(200000, self.get_tree)

    def tree_item_click(self, event):
        try:
            selection = self.tree.selection()[0]
            self.data = self.tree.item(selection, "values")
            print(self.data)
            self.item_selection = self.data[0]
            self.item_selections = self.data
            # print(self.item_selections)
        except IndexError:
            tkinter.messagebox.showinfo("单击警告", "单击结果范围异常，请重新选择！")

    def init_data(self,datas):
        result, _ = self.opr.querys(querybys=datas)
        # print(result)
        if self.opr.queryStatuss:
            return 0
        else:
            return result

    def go(self):
        datass=self.s
        # print(datass)
        self.a = int(self.C1.get())
        self.b = int(self.C2.get())
        if self.b < 10:
            self.bin=str(self.a) + "-0" + str(self.b)
        else:
            self.bin = str(self.a) + "-" + str(self.b)
        # print(self.bin)
        if self.bin is None or self.bin == '':
            tkinter.messagebox.showinfo("警告", "查询条件不能为空！")
            self.get_tree()
        else:
            re=self.opr.Ri(a=self.bin,querybys=datass[1])
            if self.opr.queryStatusr:
                tkinter.messagebox.showinfo("警告", "查询出错，请检查数据库服务是否正常")
            elif not re:
                tkinter.messagebox.showinfo("查询结果", "该查询条件没有匹配项！")
            else:
                self.list = re[0]
                # print(self.list)
                self.get_tree()
             # print(self.a,self.b)

    def shu(self):
        self.list = self.init_data(self.s)
        self.get_tree()

    def gosf(self):
        name=self.s
        num='1'
        if self.item_selection is None or self.item_selection == '':
            tkinter.messagebox.showinfo("更新警告", "未选中待更新项")
        else:
            data = [self.item_selection]
            self.opr.gos_update(updatelist=data,num=num,name=name[1])
            if self.opr.updateStatusg:
                tkinter.messagebox.showinfo("更新信息警告", "数据异常库连接异常，可能是服务关闭啦~")
            self.list = self.init_data(self.s)
            self.get_tree()

    def gose(self):
        name=self.s
        num='2'
        if self.item_selection is None or self.item_selection == '':
            tkinter.messagebox.showinfo("更新警告", "未选中待更新项")
        else:
            data = [self.item_selection]
            self.opr.gos_update(updatelist=data,num=num,name=name[1])
            if self.opr.updateStatusg:
                tkinter.messagebox.showinfo("更新信息警告", "数据异常库连接异常，可能是服务关闭啦~")
            self.list = self.init_data(self.s)
            self.get_tree()

    def gosq(self):
        name=self.s
        num='Q1'
        if self.item_selection is None or self.item_selection == '':
            tkinter.messagebox.showinfo("更新警告", "未选中待更新项")
        else:
            data = [self.item_selection]
            self.opr.gos_update(updatelist=data,num=num,name=name[1])
            if self.opr.updateStatusg:
                tkinter.messagebox.showinfo("更新信息警告", "数据异常库连接异常，可能是服务关闭啦~")
            self.list = self.init_data(self.s)
            self.get_tree()

    def gosqq(self):
        name=self.s
        num='Q2'
        if self.item_selection is None or self.item_selection == '':
            tkinter.messagebox.showinfo("更新警告", "未选中待更新项")
        else:
            data = [self.item_selection]
            self.opr.gos_update(updatelist=data,num=num,name=name[1])
            if self.opr.updateStatusg:
                tkinter.messagebox.showinfo("更新信息警告", "数据异常库连接异常，可能是服务关闭啦~")
            self.list = self.init_data(self.s)
            self.get_tree()

    def goss(self):
        name=self.s
        num='S1'
        if self.item_selection is None or self.item_selection == '':
            tkinter.messagebox.showinfo("更新警告", "未选中待更新项")
        else:
            data = [self.item_selection]
            self.opr.gos_update(updatelist=data,num=num,name=name[1])
            if self.opr.updateStatusg:
                tkinter.messagebox.showinfo("更新信息警告", "数据异常库连接异常，可能是服务关闭啦~")
            self.list = self.init_data(self.s)
            self.get_tree()

    def gosss(self):
        name=self.s
        num='S2'
        if self.item_selection is None or self.item_selection == '':
            tkinter.messagebox.showinfo("更新警告", "未选中待更新项")
        else:
            data = [self.item_selection]
            self.opr.gos_update(updatelist=data,num=num,name=name[1])
            if self.opr.updateStatusg:
                tkinter.messagebox.showinfo("更新信息警告", "数据异常库连接异常，可能是服务关闭啦~")
            self.list = self.init_data(self.s)
            self.get_tree()

    def gos0(self):
        name=self.s
        num=''
        if self.item_selection is None or self.item_selection == '':
            tkinter.messagebox.showinfo("更新警告", "未选中待更新项")
        else:
            data = [self.item_selection]
            self.opr.gos_update(updatelist=data,num=num,name=name[1])
            if self.opr.updateStatusg:
                tkinter.messagebox.showinfo("更新信息警告", "数据异常库连接异常，可能是服务关闭啦~")
            self.list = self.init_data(self.s)
            self.get_tree()

    def dao(self):
        from tkinter import messagebox
        var_box = tk.messagebox.askyesno(title='系统提示', message='是否导出')
        username = ""
        def on_click():
            global username
            username = userEntry.get().strip()
            if len(username) == 0:
                messagebox.showwarning(title='系统提示', message='请输入用户名!')
                return False
            else:
                conn = pymysql.connect(host='localhost', user='root', passwd='12345678', db='wk')
                cursor = conn.cursor()
                count = cursor.execute('select * from DATES;')
                print('has %s record' % count)
                cursor.scroll(0, mode='absolute')
                results = cursor.fetchall()
                # print(results)
                fields = cursor.description
                # print(fields)
                wbk = xlwt.Workbook()
                sheet = wbk.add_sheet('dates', cell_overwrite_ok=True)
                for ifs in range(0, len(fields)):
                    sheet.write(0, ifs, fields[ifs][0])
                ics = 1
                jcs = 0
                for ics in range(1, len(results) + 1):
                    for jcs in range(0, len(fields)):
                        sheet.write(ics, jcs, results[ics - 1][jcs])
                wbk.save(username + '.xls')
                print("完成！")
                # pa=os.path.dirname(os.path.realpath(__file__))
                pa=os.getcwd()
                # print(pa)
                #----------------------------------------------转xlsx
                sleep(0.5)
                fname = pa+'\\'+username+".xls"
                print(fname)
                excel = win32.gencache.EnsureDispatch('Excel.Application')
                wb = excel.Workbooks.Open(fname)
                wb.SaveAs(fname + "x", FileFormat=51)
                wb.Close()
                excel.Application.Quit()
                os.remove(username + '.xls')
                #-----------------------------------------------汇总
                sleep(1)
                from xlutils.copy import copy
                excel = xlrd.open_workbook(username+".xlsx")
                wb = copy(excel)
                wb.add_sheet(r'数据')
                # wb.add_sheet(r'数据2')
                count = cursor.execute('select Studentname from eng;')
                print('has %s record' % count)
                # 重置游标位置
                cursor.scroll(0, mode='absolute')
                # 搜取所有结果
                results = cursor.fetchall()
                # print(results)
                ws = wb.get_sheet(1)
                ws.write(0,0,'学生名字')
                ws.write(0,1,'1H')
                ws.write(0,2,'2H')
                ws.write(0,3,'Q1H')
                ws.write(0,4,'Q2H')
                ws.write(0,5,'S1H')
                ws.write(0,6,'S2H')
                ws.write(0,8,'总课时')
                ws.write(0,9,'剩余课时')
                ws.write(0,10,'赠送课时')
                for jcs in range(1, len(results)+1):
                    ws.write(jcs,0, results[jcs-1][0])

                for i in range(0,len(results)):
                    sql_1 = '''SELECT COUNT(*) FROM DATES WHERE  DATE REGEXP'2019' AND '''+ results[i][0]+" LIKE'1';"
                    # print(sql_1)
                    cursor.execute(sql_1)
                    result_1 = cursor.fetchall()
                    # print(result_1)
                    ws.write(i+1,1, result_1[0][0])
                for i in range(0,len(results)):
                    sql_2 = '''SELECT COUNT(*) FROM DATES WHERE  DATE REGEXP'2019' AND '''+results[i][0]+" LIKE'2';"
                    # print(sql_2)
                    cursor.execute(sql_2)
                    result_2 = cursor.fetchall()
                    # print(result_2)
                    ws.write(i+1,2, result_2[0][0])
                for i in range(0, len(results)):
                    sql_Q1 = '''SELECT COUNT(*) FROM DATES WHERE  DATE REGEXP'2019' AND '''+results[i][0]+" LIKE'Q1';"
                    # print(sql_Q1)
                    cursor.execute(sql_Q1)
                    result_Q1 = cursor.fetchall()
                    # print(result_Q1)
                    ws.write(i + 1, 3, result_Q1[0][0])
                for i in range(0, len(results)):
                    sql_Q2 = '''SELECT COUNT(*) FROM DATES WHERE  DATE REGEXP'2019' AND '''+results[i][0]+" LIKE'Q2';"
                    # print(sql_Q2)
                    cursor.execute(sql_Q2)
                    result_Q2 = cursor.fetchall()
                    # print(result_Q2)
                    ws.write(i + 1, 4, result_Q2[0][0])
                for i in range(0, len(results)):
                    sql_S1 = '''SELECT COUNT(*) FROM DATES WHERE  DATE REGEXP'2019' AND '''+results[i][0]+" LIKE'S1';"
                    # print(sql_S1)
                    cursor.execute(sql_S1)
                    result_S1 = cursor.fetchall()
                    # print(result_S1)
                    ws.write(i + 1, 5, result_S1[0][0])
                for i in range(0, len(results)):
                    sql_S2 = '''SELECT COUNT(*) FROM DATES WHERE  DATE REGEXP'2019' AND '''+results[i][0]+" LIKE'S2';"
                    # print(sql_S2)
                    cursor.execute(sql_S2)
                    result_S2 = cursor.fetchall()
                    # print(result_S2)
                    ws.write(i + 1, 6, result_S2[0][0])
                for i in range(len(results)):
                    ws.write(i+1,8,'120')
                for i in range(0,len(results)):
                    sql_Z= '''SELECT 赠送课时 from eng WHERE STUDENTNAME='''+"\""+results[i][0]+"\""+";"
                    print(sql_Z)
                    cursor.execute(sql_Z)
                    result_Z=cursor.fetchall()
                    ws.write(i+1,10,result_Z[0][0])

                wb.save(username+".xls")
                os.remove(username + '.xlsx')
                print("完成！")
                # ----------------------------------------------转xlsx
                sleep(1)
                fnames = pa + "\\" + username + ".xls"
                # print(fname)
                excel = win32.gencache.EnsureDispatch('Excel.Application')
                wb = excel.Workbooks.Open(fnames)
                wb.SaveAs(fnames + "x", FileFormat=51)
                wb.Close()
                excel.Application.Quit()
                os.remove(username + '.xls')
                print("转换完成！！")
                #------------------------------------------------数据处理
                sleep(1)
                print("数据处理开始！！")
                workbook = load_workbook(username+".xlsx")
                sheetnames = workbook.get_sheet_names()  # 获得表单名字
                # print (sheetnames)
                sheet = workbook.get_sheet_by_name(sheetnames[1])
                b = []
                all = []
                c = 0
                import copy
                for row in sheet.rows:
                    for cell in row:
                        b.append(cell.value)
                        # print(b)
                        c += 1
                        # print(c)
                        if c == 11:
                            all.append(copy.deepcopy(b))
                            # print(all)
                            c = 0
                            b.clear()
                # print(all)
                t = {}
                for i in range(1, len(all)):
                    if str(all[i][10]) == 'None':
                        zong = int(all[i][8]) - 1 * int(all[i][1]) - 2 * int(all[i][2])
                        # -1 * int(all[i][5]) - 2 * int(all[i][6])
                    else:
                        zong = int(all[i][8]) - 1 * int(all[i][1]) - 2 * int(all[i][2]) + int(all[i][10])
                        # -1 * int(all[i][5]) - 2 * int(all[i][6])
                    t[all[i][0]] = zong
                # print(t)
                for j in range(2, len(t) + 2):
                    sheet[str('J' + str(j))] = t[str(sheet[str('A' + str(j))].value)]
                workbook.save(username+".xlsx")
                print("数据处理完毕！！！")
            roots.destroy()


        def center_windowas(roots, width, height):
            screenwidth = roots.winfo_screenwidth()
            screenheight = roots.winfo_screenheight()
            size = '%dx%d+%d+%d' % (width, height, (screenwidth - width) / 2, (screenheight - height) / 2)
            roots.geometry(size)

        if var_box:
            roots = tk.Tk()
            center_windowas(roots, 300, 130)
            roots.title("数据导出")
            user = Label(roots, text='输入导出文件名字：', borderwidth=3)
            user.grid(row=2, sticky=W, padx=10, pady=25)
            userEntry = Entry(roots)
            userEntry.grid(row=2, column=1, columnspan=3, padx=2, pady=25)
            userEntry.focus_set()
            loginButton = Button(roots, text='确认名称', borderwidth=2, command=on_click)
            loginButton.grid(row=5, column=1)
            roots.mainloop()

        else:
            print("不做处理")

    def donothings(self):
        pass

    def center_windows(self, width, height):
        screenwidth = self.winfo_screenwidth()
        screenheight = self.winfo_screenheight()
        size = '%dx%d+%d+%d' % (width, height, (screenwidth - width) / 2, (screenheight - height) / 2)
        self.geometry(size)


if __name__ == '__main__':
  app = MainWindow()
  app.mainloop()